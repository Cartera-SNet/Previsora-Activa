"""
Activa IT - Descargador automático de cartas glosa (Previsora SOAT)
Versión mejorada con:
- Detener/Reiniciar
- Carpetas por IPS (forzando nombre exacto del mapa)
- Reporte Excel (descargadas + errores)
- Búsqueda flexible (Envios_D / ActaDevolucion / Carta de Objeción) con filtro correcto
- Persistencia (reanudación automática)
- Importación opcional de lista de facturas (CSV/Excel)
- Generación de ZIP parcial al detener o ante error (incluye Excel parcial y Errores)
- ZIP final incluye Excel y carpeta Errores
- API para consultar progreso y exportar a Excel
- Soporte para períodos individuales y rangos masivos
- Descarga con nombres mejorados: Factura Y [tipo_soporte]
"""

import os
import re
import json
import csv
import time
import threading
import logging
import zipfile
from datetime import datetime
from pathlib import Path
from flask import Flask, render_template, request, jsonify, send_from_directory, send_file
from io import BytesIO

# Para generar Excel
try:
    import openpyxl
    from openpyxl.styles import Font
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl no instalado. No se generará el archivo Excel.")

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__)

BASE_DIR = Path(__file__).parent
port = int(os.environ.get("PORT", 8080))
DOWNLOAD_DIR = BASE_DIR / "downloads"
DOWNLOAD_DIR.mkdir(exist_ok=True)

# ==================== MAPA DE IPS POR NIT ====================
MAPA_IPS = {
    # IPS existentes
    "900267064": "INVERSIONES_AZALUD_CLINICA_BAHIA",
    "900827065": "CENTRO_DE_DIAGNOSTICO_E_IMAGENES_BAHIA",
    "900657731": "CENTRO_MEDICO_Y_DE_REHABILITACION_BAHIA",
    "900826509": "RED_DE_URGENCIAS_DEL_MAGDALENA",
    "900513306": "FUNDACION_MARIA_REINA",
    "900600550": "INVERSIONES_MEDICAS_BARU",
    "900954800": "CENTRO_MEDICO_Y_DE_REHABILITACION_BARU",
    "900631361": "INVERSIONES_MEDICAS_VALLESALUD",
    "900257333": "ODONTOTRANS",
    "901081281": "URGETRAUMA",
    "900792417": "RED_DE_URGENCIAS_DE_LA_COSTA_PACIFICA",
    "901959993": "CLINICA_CORDIALIDAD",
    # Nuevas IPS agregadas
    "900002780": "FUNDACION_CAMPBELL",
    "901523868": "MOVID_IPS_SAS",
    "901057487": "TECNOLOGIA_DIAGNOSTICA_DEL_VALLE",
    "900558595": "FUNDACION_MEDICA_CAMPBELL",
    "901149757": "UNIDAD_MEDICA_DE_TRAUMA_VALLE_SALUD",
    "900900754": "CLINICA_VALLE_SALUD_SAN_FERNANDO",
    "900469882": "CENTRO_MEDICO_SERVISALUD_INTEGRAL_IPS_SAS",
    "802024329": "RED_DE_URGENCIA_DE_LA_COSTA_LTDA",
    "900847382": "CENTRO_MEDICO_Y_DE_REHABILITACION_VALLE_SALUD",
}

# ==================== ESTADO GLOBAL ====================
job_state = {
    "running": False,
    "stopping": False,
    "logs": [],
    "stats": {"total": 0, "descargadas": 0, "errores": 0},
    "finished": False,
    "error": None,
    "errores_detalle": [],
    "descargas_exitosas": [],
    "facturas_permitidas": [],
}
job_lock = threading.Lock()
current_browser = None
current_context = None
current_dl_dir = None
current_periodo = None
current_ips_nombre = None

# ==================== UTILIDADES DE PERÍODOS ====================
MESES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]

def validar_periodo(p):
    """Valida que un string sea un período válido (MMMYY)"""
    if not p or len(p) < 5:
        return False
    mes = p[:3]
    anio = p[3:]
    return mes in MESES and re.match(r'^\d{2}$', anio)

def generar_rango_periodos(inicio, fin):
    """Genera lista de períodos entre inicio y fin (ambos inclusive)"""
    if not validar_periodo(inicio) or not validar_periodo(fin):
        return []

    mes_inicio = MESES.index(inicio[:3])
    anio_inicio = int(inicio[3:])
    mes_fin = MESES.index(fin[:3])
    anio_fin = int(fin[3:])

    # Convertir a fecha comparable (año * 100 + mes)
    fecha_inicio = anio_inicio * 100 + mes_inicio
    fecha_fin = anio_fin * 100 + mes_fin

    if fecha_fin < fecha_inicio:
        return []

    periodos = []
    anio = anio_inicio
    mes = mes_inicio

    while True:
        anio_str = str(anio).zfill(2)
        periodos.append(MESES[mes] + anio_str)

        if anio == anio_fin and mes == mes_fin:
            break

        mes += 1
        if mes > 11:
            mes = 0
            anio += 1

    return periodos

def parse_periodo_input(periodo_input):
    """Parsea el input de período y retorna lista de períodos.
    Soporta:
    - Período único: May26
    - Rango: Dic25-May26
    """
    periodo_input = periodo_input.strip()
    if not periodo_input:
        return []

    # Detectar rango con "-"
    if '-' in periodo_input:
        parts = [p.strip() for p in periodo_input.split('-')]
        if len(parts) == 2:
            return generar_rango_periodos(parts[0], parts[1])
        return []

    # Período individual
    if validar_periodo(periodo_input):
        return [periodo_input]

    return []

# ==================== LOGGING ====================
def log(msg, level="info"):
    ts = datetime.now().strftime("%H:%M:%S")
    entry = {"ts": ts, "msg": msg, "level": level}
    with job_lock:
        job_state["logs"].append(entry)
    if level == "error":
        logger.error(msg)
    else:
        logger.info(msg)

def reset_state():
    with job_lock:
        job_state["running"] = False
        job_state["stopping"] = False
        job_state["logs"] = []
        job_state["stats"] = {"total": 0, "descargadas": 0, "errores": 0}
        job_state["finished"] = False
        job_state["error"] = None
        job_state["errores_detalle"] = []
        job_state["descargas_exitosas"] = []
        job_state["facturas_permitidas"] = []

def stop_job():
    global current_browser, current_context, current_dl_dir, current_periodo, current_ips_nombre
    with job_lock:
        job_state["stopping"] = True
    log("🛑 Solicitando detención del proceso...", "warn")
    if current_browser:
        try:
            current_browser.close()
            log("  → Navegador cerrado por solicitud de stop.")
        except Exception as e:
            log(f"  → Error al cerrar navegador: {e}", "error")
    # Generar ZIP parcial si hay archivos descargados y tenemos los datos necesarios
    generar_zip_parcial()

def generar_zip_parcial():
    """Genera un ZIP con los PDFs ya descargados hasta el momento,
       incluyendo un reporte Excel parcial y la carpeta Errores."""
    global current_dl_dir, current_periodo, current_ips_nombre
    if not current_dl_dir or not current_periodo or not current_ips_nombre:
        return
    ips_dir = current_dl_dir / current_ips_nombre
    if not ips_dir.exists():
        return

    # Obtener los datos actuales de descargas y errores
    with job_lock:
        exitosas = job_state["descargas_exitosas"].copy()
        errores = job_state["errores_detalle"].copy()

    # Generar un Excel parcial (si hay datos o si openpyxl está disponible)
    excel_parcial_path = None
    if EXCEL_AVAILABLE:
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_name = f"reporte_parcial_{timestamp}.xlsx"
            excel_parcial_path = ips_dir / excel_name
            wb = openpyxl.Workbook()
            ws_exit = wb.active
            ws_exit.title = "Descargadas"
            ws_exit.append(["N° Factura", "Estado", "IPS", "Archivo Descargado", "Fecha/Hora"])
            for ex in exitosas:
                ws_exit.append([ex.get("factura"), ex.get("estado"), current_ips_nombre, ex.get("archivo"), ex.get("timestamp")])
            ws_err = wb.create_sheet("Errores")
            ws_err.append(["N° Factura", "Estado", "IPS", "Error", "Captura pantalla", "Fecha/Hora"])
            for err in errores:
                ws_err.append([err.get("factura"), err.get("estado"), current_ips_nombre, err.get("error"), err.get("captura"), err.get("timestamp")])
            wb.save(excel_parcial_path)
            log(f"📊 Reporte Excel parcial generado: {excel_parcial_path}")
        except Exception as e:
            log(f"⚠️ No se pudo generar Excel parcial: {e}", "warn")
            excel_parcial_path = None

    # Recopilar archivos a incluir
    archivos_a_incluir = []
    # PDFs
    archivos_a_incluir.extend(ips_dir.rglob("*.pdf"))
    # Excel parcial (si existe)
    if excel_parcial_path and excel_parcial_path.exists():
        archivos_a_incluir.append(excel_parcial_path)
    # Carpeta Errores
    errores_dir = ips_dir / "Errores"
    if errores_dir.exists():
        archivos_a_incluir.extend(errores_dir.rglob("*"))

    if not archivos_a_incluir:
        return

    try:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        zip_name = f"facturas_{current_periodo}_PARCIAL_{timestamp}.zip"
        zip_path = current_dl_dir / zip_name
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for archivo in archivos_a_incluir:
                arcname = archivo.relative_to(current_dl_dir)
                zf.write(archivo, arcname=str(arcname))
        log(f"📦 ZIP parcial generado (detención/error): {zip_path}")
    except Exception as e:
        log(f"⚠️ No se pudo generar ZIP parcial: {e}", "warn")

def crear_zip_completo(dl_dir, periodo, ips_nombre):
    """Crea ZIP final incluyendo PDFs, Excel final y carpeta Errores."""
    try:
        zip_final_name = f"facturas_{periodo}.zip"
        zip_final_path = dl_dir / zip_final_name
        with zipfile.ZipFile(zip_final_path, "w", zipfile.ZIP_DEFLATED) as zf:
            ips_dir = dl_dir / ips_nombre
            if ips_dir.exists():
                # PDFs
                for pdf in ips_dir.rglob("*.pdf"):
                    zf.write(pdf, arcname=str(pdf.relative_to(dl_dir)))
                # Excel final (sin "_PARCIAL_" en el nombre)
                for excel in ips_dir.glob("reporte_*.xlsx"):
                    if "_PARCIAL_" not in excel.name:
                        zf.write(excel, arcname=str(excel.relative_to(dl_dir)))
                # Errores
                errores_dir = ips_dir / "Errores"
                if errores_dir.exists():
                    for err_file in errores_dir.rglob("*"):
                        zf.write(err_file, arcname=str(err_file.relative_to(dl_dir)))
        log(f"📦 ZIP final generado: {zip_final_path}")
        return str(zip_final_path)
    except Exception as e:
        log(f"⚠️ No se pudo generar el ZIP final: {e}", "warn")
        return None

# ==================== PERSISTENCIA (REANUDACIÓN) ====================
def cargar_progreso(ips_dir):
    progreso_path = ips_dir / "progreso.json"
    if progreso_path.exists():
        try:
            with open(progreso_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                # El progreso puede ser una lista simple o un diccionario con timestamps
                completadas = data.get("completadas", [])
                if isinstance(completadas, list):
                    return set(completadas)
                elif isinstance(completadas, dict):
                    return set(completadas.keys())
                else:
                    return set()
        except Exception as e:
            log(f"⚠️ Error al leer progreso: {e}", "warn")
    return set()

def guardar_progreso(ips_dir, completadas):
    """Guardar progreso. completadas es un set de números de factura."""
    progreso_path = ips_dir / "progreso.json"
    try:
        # Convertir el set a lista para JSON
        data = {
            "completadas": list(completadas),
            "actualizado": datetime.now().isoformat()
        }
        with open(progreso_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
    except Exception as e:
        log(f"⚠️ Error al guardar progreso: {e}", "warn")

# ==================== GENERADOR DE EXCEL ====================
def generar_reporte_excel(dl_dir, periodo, ips_nombre, exitosas, errores):
    if not EXCEL_AVAILABLE:
        return None
    excel_path = dl_dir / ips_nombre / f"reporte_{periodo}.xlsx"
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    ws_exit = wb.active
    ws_exit.title = "Descargadas"
    ws_exit.append(["N° Factura", "Estado", "IPS", "Archivo Descargado", "Fecha/Hora"])
    for ex in exitosas:
        ws_exit.append([ex.get("factura"), ex.get("estado"), ips_nombre, ex.get("archivo"), ex.get("timestamp")])

    ws_err = wb.create_sheet("Errores")
    ws_err.append(["N° Factura", "Estado", "IPS", "Error", "Captura pantalla", "Fecha/Hora"])
    for err in errores:
        ws_err.append([err.get("factura"), err.get("estado"), ips_nombre, err.get("error"), err.get("captura"), err.get("timestamp")])

    wb.save(excel_path)
    return excel_path

# ==================== FUNCIONES AUXILIARES ====================

def _find_frame_with_text(page, regex_text: str):
    js = f"() => {{ const re = new RegExp({json.dumps(regex_text)}, 'i'); return re.test(document.body?.innerText || ''); }}"
    for fr in page.frames:
        try:
            if fr.evaluate(js):
                return fr
        except:
            continue
    return None

def _cerrar_traza_factura(page):
    js = """
        () => {
            const headers = document.querySelectorAll('.ui-dialog-titlebar, .modal-header, [class*="header"]');
            for (const h of headers) {
                if (h.textContent && h.textContent.includes('Traza de Factura')) {
                    const dlg = h.closest('.ui-dialog, .modal, [role="dialog"]');
                    if (dlg) {
                        const closeBtn = dlg.querySelector('.ui-dialog-titlebar-close, button.close, [aria-label*="lose"], [class*="close"]');
                        if (closeBtn) { closeBtn.click(); return true; }
                    }
                }
            }
            return false;
        }
    """
    for fr in page.frames:
        try:
            if fr.evaluate(js):
                time.sleep(0.5)
                return
        except:
            continue

# ==================== FUNCIÓN MEJORADA DE EXTRACCIÓN DE NOMBRE DE IPS ====================
def _extraer_nombre_ips(page, target_frame, nit_usuario=None):
    """
    Extrae el nombre de la IPS forzando el uso del nombre exacto del mapa.
    Orden de prioridad:
    0. NIT extraído del nombre de usuario (ej: PREV900600550 → 900600550)
    1. Carpeta previa del mismo período (solo si existe en current_dl_dir)
    2. NIT encontrado en el HTML de la página
    3. Nombre candidato por palabras clave (con búsqueda de NIT embebido)
    4. Título de la página
    5. Fallback: IPS_DESCONOCIDA
    """
    # 0. PRIMERO: si el NIT viene del nombre de usuario, úsalo directamente
    if nit_usuario and nit_usuario in MAPA_IPS:
        nombre = MAPA_IPS[nit_usuario]
        log(f"    🏥 IPS identificada por NIT del usuario ({nit_usuario}) -> nombre del mapa: {nombre}")
        return nombre

    # 1. Si ya existe una carpeta del mismo período (current_dl_dir ya es la carpeta del período),
    #    reutilizar ese nombre para que el progreso.json se lea correctamente.
    if current_dl_dir and current_dl_dir.exists():
        for nombre_mapa in MAPA_IPS.values():
            carpeta_existente = current_dl_dir / nombre_mapa
            if carpeta_existente.exists():
                log(f"    📂 Carpeta previa encontrada, reutilizando nombre: {nombre_mapa}")
                return nombre_mapa

    def _buscar_nit_en_frame(frame):
        try:
            nit = frame.evaluate("() => { const match = document.body.innerText.match(/NIT\\s*:\\s*([\\d\\-\\s]+)/i); if(match) return match[1].replace(/[^0-9]/g, ''); return ''; }").strip()
            return nit if nit else ""
        except:
            return ""

    def _buscar_nombre_por_palabras(frame):
        keywords = ["IPS","CLINICA","HOSPITAL","CENTRO","FUNDACIÓN","URGENCIAS","SALUD","ODONTOTRANS","URGETRAUMA","CORDIALIDAD"]
        try:
            js = f"""
                () => {{
                    const keywords = {json.dumps(keywords)};
                    const elementos = document.querySelectorAll('h1, h2, h3, h4, p, div');
                    for (const el of elementos) {{
                        let txt = el.innerText.trim();
                        if (txt.length > 5 && txt.length < 100) {{
                            for (const kw of keywords) {{
                                if (txt.toUpperCase().includes(kw)) {{
                                    return txt;
                                }}
                            }}
                        }}
                    }}
                    return "";
                }}
            """
            nombre = frame.evaluate(js).strip()
            return nombre
        except:
            return ""

    # 1. Buscar NIT en todos los frames
    nit = ""
    for fr in [page] + page.frames:
        nit = _buscar_nit_en_frame(fr)
        if nit:
            log(f"    🔍 NIT encontrado en frame: {fr.name or 'principal'}")
            break

    # Si encontramos NIT y está en el mapa, usamos el nombre del mapa (sin añadir nada extra)
    if nit and nit in MAPA_IPS:
        nombre = MAPA_IPS[nit]
        log(f"    🏥 IPS identificada por NIT {nit} -> nombre forzado del mapa: {nombre}")
        return nombre

    # 2. Si no se encontró NIT o no está en el mapa, buscar por palabras clave
    nombre_candidato = ""
    for fr in [page] + page.frames:
        nombre_candidato = _buscar_nombre_por_palabras(fr)
        if nombre_candidato:
            log(f"    🔍 Nombre candidato encontrado: '{nombre_candidato}'")
            break

    if nombre_candidato:
        # Limpiar caracteres no válidos
        nombre_candidato = re.sub(r'[\\/*?:"<>|]', "", nombre_candidato).strip()
        # Buscar si dentro del nombre candidato hay un NIT conocido
        nit_embedded = re.search(r'\b(\d{9})\b', nombre_candidato)
        if nit_embedded and nit_embedded.group(1) in MAPA_IPS:
            nombre = MAPA_IPS[nit_embedded.group(1)]
            log(f"    🏥 IPS identificada por NIT embebido en texto: {nombre}")
            return nombre
        # Si no, devolver el nombre candidato limpio, pero sin añadir "_Previsora" ni similares
        nombre = re.sub(r'\s+', ' ', nombre_candidato).strip()
        log(f"    🏥 IPS identificada por texto: {nombre}")
        return nombre

    # 3. Fallback: usar el título de la página
    try:
        title = page.evaluate("() => document.title").strip()
        if title and len(title) > 5 and len(title) < 100:
            title = re.sub(r'Activa IT|BI IPS|Inteligencia de Negocio|Previsora|SOAT|Inicio', '', title, flags=re.I).strip()
            if title:
                log(f"    🏥 IPS obtenida del título: {title}")
                return title
    except:
        pass

    # 4. Fallback definitivo
    log("    ⚠️ No se pudo determinar la IPS, se usará 'IPS_DESCONOCIDA'", "warn")
    return "IPS_DESCONOCIDA"

# ==================== FUNCIÓN _download_factura (sin cambios) ====================
def _download_factura(page, context, modal_frame, fac: dict, dl_dir: Path, ips_nombre: str):
    import re
    num = fac["num"]
    tipo = fac["tipo"]

    # Determinar etiquetas según el tipo
    if tipo == "devolucion":
        target_label = "ActaDevolucion"
        target_label_norm = target_label.replace('ó', 'o').replace('í', 'i')
        subcarpeta = "Devolucion"
        nombre_soporte = "ActaDevolución"
    else:
        target_label = "Envios_D"
        target_label_norm = target_label.replace('í', 'i')
        subcarpeta = "Auditada"
        nombre_soporte = "Envios_D"

    ips_dir = dl_dir / ips_nombre
    dl_subdir = ips_dir / subcarpeta
    dl_subdir.mkdir(parents=True, exist_ok=True)

    bot_id = fac.get("botId")
    log(f"    🔗 Abriendo factura {num}...")
    num_solo_digitos = re.sub(r'\D', '', str(num))

    js_click_robusto = f"""
        () => {{
            const botId = '{bot_id}';
            const targetDigits = '{num_solo_digitos}';
            const fila = document.querySelector(`[data-bot-row-id="${{botId}}"]`);
            if (!fila) return {{ ok: false, reason: "fila_no_encontrada" }};
            fila.scrollIntoView({{block: 'center'}});
            function dispararClick(el) {{
                if (!el) return false;
                try {{ el.click(); }} catch (e) {{}}
                try {{ el.dispatchEvent(new MouseEvent('click', {{bubbles: true, cancelable: true, view: window}})); }} catch (e) {{}}
                return true;
            }}
            const candidatos = [];
            for (const a of fila.querySelectorAll('a')) {{
                const t = (a.textContent || '').trim();
                if (t.replace(/\\D/g, '') === targetDigits || candidatos.length === 0)
                    candidatos.push({{ tipo: 'a', el: a }});
            }}
            for (const el of fila.querySelectorAll('[onclick]')) {{
                if (!candidatos.find(c => c.el === el)) candidatos.push({{ tipo: 'onclick', el }});
            }}
            candidatos.push({{ tipo: 'fila', el: fila }});
            for (const td of fila.querySelectorAll('td')) candidatos.push({{ tipo: 'td', el: td }});
            for (const c of candidatos) dispararClick(c.el);
            return {{ ok: true, clickedWith: 'cascada', candidates: candidatos.length }};
        }}
    """
    result = None
    try:
        result = modal_frame.evaluate(js_click_robusto)
    except Exception as e:
        log(f"    ⚠️ Click falló: {e}", "warn")
    if not result or not result.get("ok"):
        for fr in page.frames:
            try:
                r = fr.evaluate(js_click_robusto)
                if r and r.get("ok"):
                    result = r
                    break
            except:
                continue
    if not result or not result.get("ok"):
        raise Exception(f"Click totalmente fallido para factura {num}.")
    log(f"    ✓ Click en factura {num} OK.")
    time.sleep(1.5)

    detalle_state = None
    detalle_frame = None
    for _ in range(60):
        if job_state.get("stopping"): return
        f = _find_frame_with_text(page, "Adjuntos por Factura")
        if f:
            try:
                has_traza = f.evaluate("() => /Traza de Factura/i.test(document.body?.innerText || '')")
                detalle_state = "traza" if has_traza else "adjuntos_directo"
            except:
                detalle_state = "adjuntos_directo"
            detalle_frame = f
            break
        f = _find_frame_with_text(page, "Traza de Factura")
        if f:
            detalle_state = "traza"
            detalle_frame = f
        time.sleep(0.5)
    if not detalle_frame:
        raise Exception("No apareció 'Traza de Factura' ni 'Adjuntos por Factura'.")
    time.sleep(1.5)
    log(f"    ✅ Detalle abierto (modo: {detalle_state}).")

    if detalle_state == "traza":
        log("    📑 Forzando cambio a pestaña 'Soportes'...")
        soportes_ok = False
        for intento in range(5):
            if job_state.get("stopping"): return
            for fr in page.frames:
                try:
                    has_tabs = fr.evaluate(r"""() => {
                        const txt = (document.body?.innerText || '').replace(/\n/g, ' ');
                        return /Factura.*Detalles.*Soportes/i.test(txt);
                    }""")
                    if has_tabs:
                        try:
                            fr.locator("text=Soportes").first.click(timeout=5000)
                            soportes_ok = True
                            break
                        except:
                            clicked = fr.evaluate("""() => {
                                for (const el of document.querySelectorAll('*')) {
                                    if ((el.textContent||'').trim() === 'Soportes') {
                                        el.click(); return true;
                                    }
                                }
                                return false;
                            }""")
                            if clicked:
                                soportes_ok = True
                                break
                except:
                    continue
            if soportes_ok:
                break
            time.sleep(1)
        if not soportes_ok:
            log("    ⚠️ No se pudo clickear Soportes", "warn")
        else:
            time.sleep(3)

    log("    ⏳ Esperando 'Adjuntos por Factura'...")
    adjuntos_frame = None
    for _ in range(90):
        if job_state.get("stopping"): return
        for fr in page.frames:
            try:
                if fr.evaluate("() => /Adjuntos por Factura|Buscar por.*Fecha/i.test(document.body?.innerText || '')"):
                    adjuntos_frame = fr
                    break
            except:
                continue
        if adjuntos_frame:
            break
        time.sleep(0.5)
    if not adjuntos_frame:
        raise Exception("No se encontró sección 'Adjuntos por Factura'.")
    for _ in range(35):
        if job_state.get("stopping"): return
        try:
            busy = adjuntos_frame.evaluate("() => /Procesando Solicitud/i.test(document.body?.innerText || '')")
            if not busy:
                break
        except:
            pass
        time.sleep(1)
    time.sleep(1)
    log("    ✅ Adjuntos cargados.")

    search_frame = adjuntos_frame

    # Función mejorada para escribir en el buscador y disparar la lupa
    def _escribir_buscador(texto):
        # Limpiar input antes
        search_frame.evaluate("""
            () => {
                const inputs = document.querySelectorAll('input');
                for (const input of inputs) {
                    const ph = (input.placeholder || '').toLowerCase();
                    if (ph.includes('buscar') || ph.includes('filtrar') || ph.includes('nombre')) {
                        input.value = '';
                        input.dispatchEvent(new Event('input', { bubbles: true }));
                        break;
                    }
                }
            }
        """)
        time.sleep(0.5)
        # Escribir el texto y disparar eventos
        search_frame.evaluate(f"""
            () => {{
                const target = '{texto.replace('í', 'i')}';
                const inputs = document.querySelectorAll('input');
                let searchInput = null;
                for (const input of inputs) {{
                    const ph = (input.placeholder || '').toLowerCase();
                    if (ph.includes('buscar') || ph.includes('filtrar') || ph.includes('nombre')) {{
                        searchInput = input;
                        break;
                    }}
                }}
                if (!searchInput) return;
                searchInput.focus();
                searchInput.select();
                const nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                if (nativeSetter) nativeSetter.call(searchInput, target);
                else searchInput.value = target;
                searchInput.dispatchEvent(new Event('input', {{ bubbles: true }}));
                searchInput.dispatchEvent(new Event('change', {{ bubbles: true }}));
                // Buscar botón de lupa
                let parent = searchInput.closest('div, td, form, span');
                if (parent) {{
                    const btns = parent.querySelectorAll('button, a, [role="button"], span');
                    for (const btn of btns) {{
                        const html = (btn.outerHTML || '').toLowerCase();
                        const title = (btn.title || '').toLowerCase();
                        if (html.includes('search') || html.includes('lup') || title.includes('search')) {{
                            btn.click();
                            return;
                        }}
                    }}
                }}
                const svgs = document.querySelectorAll('svg');
                for (const svg of svgs) {{
                    if ((svg.outerHTML || '').toLowerCase().includes('search')) {{
                        const container = svg.closest('button, a, [role="button"]');
                        if (container) {{ container.click(); return; }}
                    }}
                }}
                // Fallback: presionar Enter
                searchInput.dispatchEvent(new KeyboardEvent('keypress', {{ key: 'Enter', bubbles: true }}));
            }}
        """)
        time.sleep(2)
        # Esperar a que termine "Procesando Solicitud"
        for _ in range(40):
            if job_state.get("stopping"): return
            processing = False
            for fr in page.frames:
                try:
                    if fr.evaluate("() => /Procesando Solicitud/i.test(document.body?.innerText || '')"):
                        processing = True
                        break
                except:
                    pass
            if not processing:
                break
            time.sleep(0.5)
        time.sleep(2)

    # ---------- BÚSQUEDA DE SOPORTES CON MEJORAS ----------
    # Intentar primero con la etiqueta principal (Envios_D o ActaDevolucion)
    log(f"    🔍 Buscando '{target_label}'...")
    _escribir_buscador(target_label)
    archivo_seleccionado = False
    tipo_encontrado = None  # Rastrear qué tipo de soporte se encontró
    posibles_nombres = list({target_label, target_label_norm})

    for intento in range(4):
        if job_state.get("stopping"): return
        for fr in page.frames:
            try:
                resultado = fr.evaluate(f"""
                    () => {{
                        const nombres = {json.dumps(posibles_nombres)};
                        let contenedor = null;
                        const elementos = document.querySelectorAll('td, div, span, li, p, tr');
                        for (const el of elementos) {{
                            const txt = (el.innerText || '').trim();
                            for (const nombre of nombres) {{
                                if (txt === nombre) {{
                                    contenedor = el.closest('div[class*="file"], li[class*="file"], tr, div[class*="item"], div[class*="attach"], div[class*="row"]');
                                    if (!contenedor) contenedor = el.closest('div, li, tr');
                                    break;
                                }}
                            }}
                            if (contenedor) break;
                        }}
                        if (!contenedor) return {{ ok: false }};
                        let check = contenedor.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                        if (!check) check = contenedor.parentElement?.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                        if (check) {{
                            if (!check.checked) {{
                                check.click();
                                check.checked = true;
                                check.dispatchEvent(new Event('change', {{ bubbles: true }}));
                            }}
                            return {{ ok: true, metodo: 'checkbox' }};
                        }}
                        let iconoPdf = null;
                        const candidatosPdf = contenedor.querySelectorAll('img, svg, i, div');
                        for (const el of candidatosPdf) {{
                            const src = el.getAttribute('src') || '';
                            const lbl = el.getAttribute('aria-label') || '';
                            const cls = el.className || '';
                            if (src.toLowerCase().includes('pdf') || lbl.toLowerCase().includes('pdf') ||
                                cls.toLowerCase().includes('pdf') || cls.toLowerCase().includes('file')) {{
                                iconoPdf = el; break;
                            }}
                        }}
                        if (iconoPdf) {{
                            iconoPdf.click();
                            return {{ ok: true, metodo: 'icono_pdf' }};
                        }}
                        contenedor.click();
                        contenedor.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true }}));
                        contenedor.dispatchEvent(new MouseEvent('dblclick', {{ bubbles: true, cancelable: true }}));
                        return {{ ok: true, metodo: 'contenedor_forzado' }};
                    }}
                """)
                if resultado and resultado.get('ok'):
                    log(f"    ✅ Selección realizada (método: {resultado.get('metodo')})")
                    archivo_seleccionado = True
                    tipo_encontrado = nombre_soporte  # Guardar el tipo encontrado
                    break
            except Exception as e:
                log(f"    ⚠️ Error en intento {intento+1}: {e}", "warn")
        if archivo_seleccionado:
            break
        log(f"    🔄 Reintentando selección ({intento+1}/4)...")
        time.sleep(2)

    # ---------- SEGUNDA BÚSQUEDA: CARTA DE OBJECIÓN (si no se encontró Envios_D/ActaDevolucion) ----------
    if not archivo_seleccionado:
        log(f"    ⚠️ No se encontró '{target_label}'. Intentando con 'Carta de'...")
        texto_busqueda = "Carta de"
        _escribir_buscador(texto_busqueda)

        archivo_seleccionado = False
        for intento in range(4):
            if job_state.get("stopping"): return
            for fr in page.frames:
                try:
                    resultado = fr.evaluate(f"""
                        () => {{
                            const buscarTexto = '{texto_busqueda}';
                            function normalizar(s) {{
                                return s.toLowerCase().normalize("NFD").replace(/[\\u0300-\\u036f]/g, "");
                            }}
                            const elementos = document.querySelectorAll('td, div, span, li, p, tr');
                            for (const el of elementos) {{
                                const txt = (el.innerText || '').trim();
                                if (normalizar(txt).includes(normalizar(buscarTexto))) {{
                                    let contenedor = el.closest('div[class*="file"], li[class*="file"], tr, div[class*="item"], div[class*="attach"], div[class*="row"]');
                                    if (!contenedor) contenedor = el.closest('div, li, tr');
                                    if (contenedor) {{
                                        let check = contenedor.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                                        if (!check) check = contenedor.parentElement?.querySelector('input[type="checkbox"], input[type="radio"], [role="checkbox"]');
                                        if (check) {{
                                            if (!check.checked) {{
                                                check.click();
                                                check.checked = true;
                                                check.dispatchEvent(new Event('change', {{ bubbles: true }}));
                                            }}
                                            return {{ ok: true, metodo: 'checkbox', texto: txt }};
                                        }}
                                        let iconoPdf = null;
                                        const candidatosPdf2 = contenedor.querySelectorAll('img, svg, i, div');
                                        for (const el of candidatosPdf2) {{
                                            const src = el.getAttribute('src') || '';
                                            const lbl = el.getAttribute('aria-label') || '';
                                            const cls = el.className || '';
                                            if (src.toLowerCase().includes('pdf') || lbl.toLowerCase().includes('pdf') ||
                                                cls.toLowerCase().includes('pdf')) {{
                                                iconoPdf = el; break;
                                            }}
                                        }}
                                        if (iconoPdf) {{
                                            iconoPdf.click();
                                            return {{ ok: true, metodo: 'icono_pdf', texto: txt }};
                                        }}
                                        contenedor.click();
                                        contenedor.dispatchEvent(new MouseEvent('click', {{ bubbles: true, cancelable: true }}));
                                        contenedor.dispatchEvent(new MouseEvent('dblclick', {{ bubbles: true, cancelable: true }}));
                                        return {{ ok: true, metodo: 'contenedor_forzado', texto: txt }};
                                    }}
                                }}
                            }}
                            return {{ ok: false }};
                        }}
                    """)
                    if resultado and resultado.get('ok'):
                        log(f"    ✅ Selección realizada con '{texto_busqueda}' (método: {resultado.get('metodo')}) - Texto encontrado: '{resultado.get('texto')}'")
                        archivo_seleccionado = True
                        tipo_encontrado = "Carta de Objecion"  # Guardar el tipo encontrado
                        break
                except Exception as e:
                    log(f"    ⚠️ Error en intento {intento+1} para '{texto_busqueda}': {e}", "warn")
            if archivo_seleccionado:
                break
            log(f"    🔄 Reintentando '{texto_busqueda}' ({intento+1}/4)...")
            time.sleep(2)

    if not archivo_seleccionado:
        raise Exception(f"No se pudo seleccionar el archivo (intentó '{target_label}' y 'Carta de')")

    # Confirmar que no haya mensaje de error "Debe seleccionar..."
    log("    ⏳ Esperando confirmación de selección...")
    for _ in range(20):
        if job_state.get("stopping"): return
        hay_error = False
        for fr in page.frames:
            try:
                if fr.evaluate("() => /Debe seleccionar por lo menos un documento/i.test(document.body?.innerText || '')"):
                    hay_error = True
                    break
            except:
                pass
        if not hay_error:
            log("    ✅ Selección confirmada")
            break
        time.sleep(1)

    # ---------- ABRIR DOCUMENTO ----------
    log(f"    👁️ Buscando botón 'Abrir Documento'...")
    pdf_data = None
    pdf_url = None

    boton_encontrado = False
    start_time = time.time()
    while time.time() - start_time < 15:
        if job_state.get("stopping"): return
        for fr in page.frames:
            try:
                btn = fr.locator('button[title="Abrir Documento"], button[aria-label="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                if btn.is_visible(timeout=2000):
                    boton_encontrado = True
                    break
            except:
                pass
        if boton_encontrado:
            break
        time.sleep(0.5)
    else:
        raise Exception("Botón 'Abrir Documento' no encontrado")

    for reintento in range(2):
        if job_state.get("stopping"): return
        new_page = None
        try:
            with context.expect_page(timeout=30000) as page_info:
                for fr in page.frames:
                    try:
                        btn = fr.locator('button[title="Abrir Documento"], button[aria-label="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                        if btn.is_visible(timeout=5000):
                            for _ in range(10):
                                if btn.is_enabled():
                                    break
                                time.sleep(0.5)
                            btn.click()
                            log("    ✅ Clic en botón 'Abrir Documento'")
                            break
                    except:
                        pass
            new_page = page_info.value
            for _ in range(30):
                if job_state.get("stopping"): return
                url = new_page.url
                if url and url != "about:blank" and ("amazonaws" in url or ".pdf" in url.lower()):
                    pdf_url = url
                    break
                time.sleep(0.5)
        except Exception as e:
            log(f"    ⚠️ Intento {reintento+1}: No se abrió nueva pestaña: {e}", "warn")
        finally:
            if new_page:
                try:
                    new_page.close()
                except:
                    pass

        if pdf_url:
            try:
                response = context.request.get(pdf_url, timeout=60000)
                if response.ok:
                    pdf_data = response.body()
                    log(f"    ✅ PDF descargado ({len(pdf_data)//1024} KB)")
                    break
            except Exception as e:
                log(f"    ⚠️ Error descargando: {e}", "warn")

        if not pdf_data:
            log("    ⏳ Intentando descarga directa...")
            try:
                with page.expect_download(timeout=30000) as download_info:
                    for fr in page.frames:
                        try:
                            btn = fr.locator('button[title="Abrir Documento"], button:has(i.fa-eye), button:has(i.bi-eye)').first
                            if btn.is_visible(timeout=3000):
                                btn.click()
                                break
                        except:
                            pass
                download = download_info.value
                pdf_data = download.path().read_bytes() if download.path() else None
                log("    ✅ Descarga directa capturada")
                break
            except Exception as e:
                log(f"    ⚠️ No se capturó descarga: {e}", "warn")

        if not pdf_data:
            log(f"    🔄 Reintento {reintento+1}/2...")
            time.sleep(2)

    if not pdf_data:
        raise Exception("No se pudo obtener el PDF")

    # ---------- MEJORAR NOMBRE DEL ARCHIVO ----------
    # Determinar el tipo de soporte encontrado para el nombre del archivo
    soporte_encontrado = tipo_encontrado if tipo_encontrado else nombre_soporte

    # El nombre del archivo será: {num}_{soporte_encontrado}.pdf
    safe_name = re.sub(r"[^\w\-_.]", "_", f"{num}_{soporte_encontrado}.pdf")
    out_path = dl_subdir / safe_name
    out_path.write_bytes(pdf_data)
    log(f"    💾 PDF guardado: {out_path.name} ({len(pdf_data)//1024} KB)")

    with job_lock:
        job_state["descargas_exitosas"].append({
            "factura": num,
            "estado": fac["estado"],
            "archivo": str(out_path),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })

    _cerrar_traza_factura(page)
    time.sleep(0.8)

# ==================== AUTOMATIZACIÓN PRINCIPAL ====================
# (El resto del código de run_automation y rutas Flask es idéntico al original,
#  no se modifica nada más. Para ahorrar espacio, se incluye tal cual estaba,
#  pero asegurando que la función _extraer_nombre_ips es la nueva.)

def run_automation(usuario: str, password: str, periodo: str, download_path: str):
    from playwright.sync_api import sync_playwright
    global current_browser, current_context, current_dl_dir, current_periodo, current_ips_nombre

    dl_dir = Path(download_path)
    dl_dir.mkdir(parents=True, exist_ok=True)
    ips_nombre_actual = "IPS_SIN_NOMBRE"
    zip_parcial_generado = False

    current_dl_dir = dl_dir
    current_periodo = periodo

    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(accept_downloads=True, viewport={"width": 1500, "height": 900})
            page = context.new_page()
            current_browser = browser
            current_context = context

            log("🔐 Iniciando sesión en Activa IT...")
            if job_state.get("stopping"): return
            page.goto("https://activa-it.net/Login.aspx", wait_until="networkidle", timeout=60000)
            log(f"  → Usuario: {usuario}")
            page.fill('input[placeholder="Usuario"]', usuario)
            page.fill('input[placeholder="Contraseña"]', password)
            try:
                checkbox = page.locator('input[type="checkbox"]').first
                if not checkbox.is_checked():
                    checkbox.check()
            except:
                pass
            page.click('button:has-text("Inicio de sesión"), input[value="Inicio de sesión"]')
            page.wait_for_url("**/Index.aspx", timeout=60000)
            time.sleep(2)
            log("✅ Sesión iniciada correctamente.")
            if job_state.get("stopping"): return

            log("📂 Navegando a módulo BI IPS...")
            time.sleep(3)

            def _find_periodo_in_frames():
                js_check = f"""
                    () => {{
                        const bodyText = (document.body?.innerText || '').toLowerCase();
                        const periodo = '{periodo}'.toLowerCase();
                        if (bodyText.includes(periodo)) return true;
                        const variaciones = ['abr26', 'abr-26', 'abr.26', 'abr/26', 'abr2026'];
                        return variaciones.some(v => bodyText.includes(v));
                    }}
                """
                for fr in page.frames:
                    try:
                        if fr.evaluate(js_check):
                            return fr
                    except:
                        continue
                return None

            if job_state.get("stopping"): return
            clicked = False
            for intento in range(3):
                try:
                    page.locator("text=BI IPS").first.click(timeout=15000)
                    clicked = True
                    log("  ✓ Click directo en 'BI IPS' OK.")
                    break
                except:
                    pass
                try:
                    page.click("text=Inteligencia de Negocio", timeout=8000)
                    time.sleep(1)
                    page.click("text=BI IPS", timeout=8000)
                    clicked = True
                    log("  ✓ Click vía 'Inteligencia de Negocio' + 'BI IPS' OK.")
                    break
                except:
                    pass
                try:
                    page.click("[class*='menu-toggle'], [class*='hamburger'], .sidebar-toggle", timeout=5000)
                    time.sleep(2)
                    page.click("text=BI IPS", timeout=8000)
                    clicked = True
                    log("  ✓ Click vía hamburguesa + 'BI IPS' OK.")
                    break
                except Exception as e:
                    log(f"    ⚠️ Intento {intento+1} falló: {e}", "warn")
                    time.sleep(2)
            if not clicked:
                raise Exception("No se encontró el módulo BI IPS en el menú.")

            time.sleep(3)
            log("✅ Módulo BI IPS abierto. Buscando período...")
            target_frame = None
            for i in range(120):
                if job_state.get("stopping"): return
                target_frame = _find_periodo_in_frames()
                if target_frame:
                    log(f"✅ Período '{periodo}' detectado tras {(i+1)*0.5:.1f}s.")
                    break
                time.sleep(0.5)
            if not target_frame:
                raise Exception(f"No se pudo localizar el período '{periodo}' tras 60s.")

            log("🏥 Obteniendo nombre de la IPS...")
            # Intentar extraer NIT del nombre de usuario (ej: PREV900600550 → 900600550)
            nit_from_usuario = re.search(r'(\d{9,12})', usuario)
            nit_from_usuario = nit_from_usuario.group(1) if nit_from_usuario else None
            if nit_from_usuario:
                log(f"    🔑 NIT extraído del usuario '{usuario}': {nit_from_usuario}")
            ips_nombre_actual = _extraer_nombre_ips(page, target_frame, nit_usuario=nit_from_usuario)
            current_ips_nombre = ips_nombre_actual

            if job_state.get("stopping"): return
            log(f"📅 Click en columna Cant del período '{periodo}'...")
            click_result = target_frame.evaluate(f"""
                () => {{
                    const rows = document.querySelectorAll('tr');
                    for (const row of rows) {{
                        const cells = row.querySelectorAll('td');
                        if (cells.length < 3) continue;
                        const firstText = cells[0].textContent.trim();
                        if (firstText !== '{periodo}') continue;
                        const links = row.querySelectorAll('a');
                        if (links.length === 0) return {{ ok: false, reason: 'sin_links' }};
                        const firstLink = links[0];
                        const value = firstLink.textContent.trim();
                        if (value === '0') return {{ ok: false, reason: 'cant_cero', value: '0' }};
                        firstLink.scrollIntoView({{block: 'center'}});
                        firstLink.click();
                        return {{ ok: true, value: value }};
                    }}
                    return {{ ok: false, reason: 'fila_no_encontrada' }};
                }}
            """)
            if click_result.get("reason") == "cant_cero":
                log(f"ℹ️ El período '{periodo}' tiene 0 facturas radicadas.", "warn")
                browser.close()
                return
            if not click_result.get("ok"):
                raise Exception(f"No se pudo hacer click en Cant de '{periodo}': {click_result.get('reason')}")
            log(f"  → Click en Cant: {click_result.get('value')}")

            log("⏳ Esperando modal 'Listado de facturas recibidas'...")
            modal_frame = None
            for _ in range(60):
                if job_state.get("stopping"): return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Listado de facturas recibidas/i.test(document.body?.innerText || '')"):
                            modal_frame = fr
                            break
                    except:
                        continue
                if modal_frame:
                    break
                time.sleep(0.5)
            if not modal_frame:
                raise Exception("El modal 'Listado de facturas recibidas' no apareció.")

            log("⏳ Esperando datos del listado...")
            data_frame = None
            tiempo_espera = 0
            while tiempo_espera < 60:
                if job_state.get("stopping"): return
                for fr in page.frames:
                    try:
                        if fr.evaluate("() => /Pendiente de recibir Informaci|Devoluci[oó]n de entrada/i.test(document.body?.innerText || '')"):
                            data_frame = fr
                            break
                    except:
                        continue
                if data_frame:
                    break
                time.sleep(0.5)
                tiempo_espera += 0.5

            if not data_frame:
                log("⚠️ No se encontraron facturas con los estados objetivo.", "warn")
                browser.close()
                return

            log(f"✅ Datos detectados en frame '{data_frame.name or '(main)'}'.")
            time.sleep(2)

            log("🔍 Extrayendo facturas...")
            js_extract = r"""
            (state) => {
                const ESTADOS = [
                    { nombre: 'Auditada: Pendiente de recibir Informacion', regex: /auditada\s*:\s*pendiente\s+de\s+recibir\s+informaci[oó]n/i, tipo: 'auditada' },
                    { nombre: 'En radicacion: Devolución de entrada', regex: /en\s+radicaci[oó]n\s*:\s*devoluci[oó]n\s+de\s+entrada/i, tipo: 'devolucion' },
                    { nombre: 'En auditoria: Pendiente de informar Orden de pago al Pagador', regex: /en\s+auditori?a\s*:\s*pendiente\s+de\s+informar\s+orden\s+de\s+pago\s+al\s+pagador/i, tipo: 'auditada' },
                ];
                const filas = document.querySelectorAll('tr, [role="row"], li');
                const nuevas = [];
                for (const fila of filas) {
                    const fullText = (fila.innerText || '').replace(/\s+/g, ' ').trim();
                    if (!fullText || fullText.length < 20 || fullText.length > 400) continue;
                    if (!/\d{2}\/\d{2}\/\d{4}/.test(fullText)) continue;
                    let tipoDetectado = null, nombreEstado = null;
                    for (const e of ESTADOS) {
                        if (e.regex.test(fullText)) { tipoDetectado = e.tipo; nombreEstado = e.nombre; break; }
                    }
                    if (!tipoDetectado) continue;
                    const tokens = fullText.split(/\s+/);
                    const candidatosNum = tokens.filter(t => { const digits = t.replace(/\D/g, ''); return digits.length >= 6 && digits.length <= 10; });
                    if (candidatosNum.length === 0 || candidatosNum.length > 6) continue;
                    const numNorm = candidatosNum[0].replace(/\D/g, '');
                    if (state.seen.includes(numNorm)) continue;
                    const botId = 'bot_' + state.nextId;
                    state.nextId++;
                    fila.setAttribute('data-bot-row-id', botId);
                    nuevas.push({
                        botId: botId, num: numNorm, rawNum: candidatosNum[0],
                        tipo: tipoDetectado, estado: nombreEstado,
                        textoFila: fullText.slice(0, 150), tagName: fila.tagName.toLowerCase(),
                    });
                    state.seen.push(numNorm);
                }
                return { nuevas: nuevas, total: state.seen.length };
            }
            """
            extract_state = {"nextId": 0, "seen": []}
            facturas_acumuladas = []
            rondas_sin_nuevos = 0
            for ronda in range(20):
                if job_state.get("stopping"): return
                try:
                    res = data_frame.evaluate(js_extract, extract_state)
                except:
                    res = {"nuevas": []}
                nuevas = res.get("nuevas", [])
                if nuevas:
                    facturas_acumuladas.extend(nuevas)
                    rondas_sin_nuevos = 0
                    log(f"  Ronda {ronda+1}: +{len(nuevas)} (Total: {len(facturas_acumuladas)})")
                else:
                    rondas_sin_nuevos += 1
                extract_state["seen"] = list(set(extract_state["seen"] + [n["num"] for n in nuevas]))
                if rondas_sin_nuevos >= 5:
                    break
                try:
                    data_frame.evaluate("() => { const scrollables = document.querySelectorAll('div, table, tbody, [class*=\"scroll\"]'); for (const s of scrollables) { if (s.scrollHeight > s.clientHeight + 20) s.scrollTop += s.clientHeight * 0.8; } window.scrollBy(0, window.innerHeight * 0.8); }")
                except:
                    pass
                time.sleep(0.5)
            log(f"📊 {len(facturas_acumuladas)} facturas detectadas.")
            facturas_objetivo = facturas_acumuladas

            # ========== PERSISTENCIA Y FILTRO ==========
            ips_dir = dl_dir / ips_nombre_actual
            completadas = cargar_progreso(ips_dir)

            # Filtrar facturas ya descargadas
            facturas_pendientes = []
            for fac in facturas_objetivo:
                if fac['num'] in completadas:
                    log(f"⏭️ Factura {fac['num']} ya descargada en ejecución anterior, omitiendo.")
                    with job_lock:
                        job_state["stats"]["descargadas"] += 1
                        job_state["descargas_exitosas"].append({
                            "factura": fac['num'],
                            "estado": fac['estado'],
                            "archivo": str(ips_dir / ("Auditada" if fac['tipo']=='auditada' else "Devolucion") / f"Factura_{fac['num']}_{('Envios_D' if fac['tipo']=='auditada' else 'ActaDevolucion')}.pdf"),
                            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        })
                else:
                    facturas_pendientes.append(fac)

            # Aplicar filtro opcional por lista de facturas permitidas
            with job_lock:
                permitidas = job_state.get("facturas_permitidas", [])
            if permitidas:
                original_count = len(facturas_pendientes)
                facturas_pendientes = [fac for fac in facturas_pendientes if fac['num'] in permitidas]
                log(f"📋 Filtro activo: solo {len(facturas_pendientes)} de {original_count} facturas están en la lista permitida.")

            log(f"📋 Facturas pendientes por procesar en esta ejecución: {len(facturas_pendientes)}")

            # Actualizar estadísticas totales
            with job_lock:
                job_state["stats"]["total"] = len(facturas_pendientes) + job_state["stats"]["descargadas"]
                job_state["stats"]["errores"] = 0

            cnt_aud = sum(1 for f in facturas_pendientes if f["tipo"] == "auditada")
            cnt_dev = sum(1 for f in facturas_pendientes if f["tipo"] == "devolucion")
            log("📋 RESUMEN DE FACTURAS PENDIENTES:")
            log(f"  • Auditada: {cnt_aud}")
            log(f"  • Devolucion: {cnt_dev}")
            log(f"  TOTAL: {len(facturas_pendientes)}")
            if not facturas_pendientes:
                log("ℹ️ No hay facturas pendientes por procesar.")
                browser.close()
                with job_lock:
                    exitosas = job_state["descargas_exitosas"].copy()
                    errores = job_state["errores_detalle"].copy()
                generar_reporte_excel(dl_dir, periodo, ips_nombre_actual, exitosas, errores)
                crear_zip_completo(dl_dir, periodo, ips_nombre_actual)
                return

            # ========== PROCESAR FACTURAS PENDIENTES ==========
            # ── Función interna: procesar una lista de facturas en la sesión activa ──
            def _procesar_lista(lista, intento_num):
                fallidas = []
                for idx, fac in enumerate(lista, 1):
                    if job_state.get("stopping"):
                        log("🛑 Proceso detenido por el usuario.")
                        if not zip_parcial_generado:
                            generar_zip_parcial()
                        return None  # señal de detención
                    log(f"[Intento {intento_num}][{idx}/{len(lista)}] Factura {fac['num']} ({fac['tipo']})...")
                    try:
                        _download_factura(page, context, data_frame, fac, dl_dir, ips_nombre_actual)
                        with job_lock:
                            job_state["stats"]["descargadas"] += 1
                            job_state["stats"]["errores"] = max(0, job_state["stats"]["errores"] - 1) if intento_num > 1 else job_state["stats"]["errores"]
                        completadas.add(fac['num'])
                        guardar_progreso(ips_dir, completadas)
                        log(f"  ✅ Descargada: {fac['num']}", "success")
                    except Exception as e:
                        with job_lock:
                            if intento_num == 1:
                                job_state["stats"]["errores"] += 1
                            error_msg = str(e)
                            if "No se pudo seleccionar el archivo" in error_msg:
                                if fac['tipo'] == 'auditada':
                                    error_msg = f"En la factura {fac['num']} no se encontró soporte Envios_D ni Carta de Objecion"
                                else:
                                    error_msg = f"En la factura {fac['num']} no se encontró soporte ActaDevolucion ni Carta de Objecion"
                        log(f"  ⚠️ Error intento {intento_num}: {error_msg}", "error")
                        _cerrar_traza_factura(page)
                        time.sleep(1)
                        fallidas.append(fac)
                return fallidas

            # ── Primer pase ──
            MAX_REINTENTOS = 5
            fallidas = _procesar_lista(facturas_pendientes, 1)

            # ── Reintentos automáticos ──
            if fallidas is None:  # usuario detuvo
                browser.close()
                return

            intento = 2
            while fallidas and intento <= MAX_REINTENTOS and not job_state.get("stopping"):
                log(f"🔄 {len(fallidas)} factura(s) con error. Reintentando automáticamente ({intento}/{MAX_REINTENTOS})...", "warn")
                time.sleep(3)
                # Limpiar errores anteriores de esas facturas para no duplicar en el reporte
                nums_fallidas = {f['num'] for f in fallidas}
                with job_lock:
                    job_state["errores_detalle"] = [e for e in job_state["errores_detalle"] if e["factura"] not in nums_fallidas]
                fallidas_nuevo = _procesar_lista(fallidas, intento)
                if fallidas_nuevo is None:
                    browser.close()
                    return
                fallidas = fallidas_nuevo
                intento += 1

            # ── Registrar errores persistentes al final ──
            if fallidas:
                log(f"⛔ {len(fallidas)} factura(s) no pudieron descargarse tras {MAX_REINTENTOS} intentos.", "error")
                with job_lock:
                    for fac in fallidas:
                        nums_existentes = {e["factura"] for e in job_state["errores_detalle"]}
                        if fac['num'] not in nums_existentes:
                            error_info = {
                                "factura": fac['num'],
                                "estado": fac['estado'],
                                "error": f"Error persistente tras {MAX_REINTENTOS} intentos automáticos",
                                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                                "captura": ""
                            }
                            try:
                                errores_dir = ips_dir / "Errores"
                                errores_dir.mkdir(parents=True, exist_ok=True)
                                cap_path = errores_dir / f"ERROR_PERSISTENTE_{fac['num']}.png"
                                page.screenshot(path=str(cap_path))
                                error_info["captura"] = str(cap_path)
                            except:
                                pass
                            job_state["errores_detalle"].append(error_info)
                    job_state["stats"]["errores"] = len(job_state["errores_detalle"])

            browser.close()

            # ========== GENERAR EXCEL FINAL ==========
            with job_lock:
                exitosas = job_state["descargas_exitosas"].copy()
                errores = job_state["errores_detalle"].copy()
            excel_path = generar_reporte_excel(dl_dir, periodo, ips_nombre_actual, exitosas, errores)
            if excel_path:
                log(f"📊 Reporte Excel generado: {excel_path}")
            else:
                log("⚠️ No se pudo generar el Excel (openpyxl no instalado o error).", "warn")

            # ========== ZIP FINAL (incluye Excel y Errores) ==========
            crear_zip_completo(dl_dir, periodo, ips_nombre_actual)

            if fallidas:
                log(f"⚠️ Proceso completado con {len(fallidas)} factura(s) con error persistente. Ver Excel para detalle.")
            else:
                log("🎉 Proceso completado sin errores.")

    except Exception as e:
        if not job_state.get("stopping"):
            log(f"💥 Error crítico: {e}", "error")
            with job_lock:
                job_state["error"] = str(e)
        else:
            log("Proceso detenido por el usuario.")
        if not zip_parcial_generado:
            generar_zip_parcial()
    finally:
        with job_lock:
            job_state["running"] = False
            job_state["finished"] = True
            job_state["stopping"] = False
        current_browser = None
        current_context = None
        current_dl_dir = None
        current_periodo = None
        current_ips_nombre = None

# ==================== RUTAS FLASK ====================
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/api/start", methods=["POST"])
def start_job():
    data = request.json or {}
    usuario = data.get("usuario", "").strip()
    password = data.get("password", "").strip()
    periodo_input = data.get("periodo", "").strip()
    custom_path = data.get("download_path", "").strip()

    if not all([usuario, password, periodo_input]):
        return jsonify({"ok": False, "error": "Faltan campos requeridos"}), 400

    periodos = parse_periodo_input(periodo_input)
    if not periodos:
        return jsonify({"ok": False, "error": f"Formato de período inválido: '{periodo_input}'. Use MMMYY (ej: May26) o rango MMMYY-MMMYY"}), 400

    with job_lock:
        if job_state["running"]:
            return jsonify({"ok": False, "error": "Ya hay un proceso en ejecución"}), 409
        job_state["running"] = True
        job_state["finished"] = False
        job_state["error"] = None
        job_state["stats"] = {"total": 0, "descargadas": 0, "errores": 0}
        job_state["errores_detalle"] = []
        job_state["descargas_exitosas"] = []

    dl_path = custom_path if custom_path else str(DOWNLOAD_DIR / periodo_input)
    periodo_principal = periodos[0] if len(periodos) == 1 else periodo_input

    if len(periodos) > 1:
        log(f"📅 Procesando rango de {len(periodos)} períodos: {periodos[0]} → {periodos[-1]}")
        job_state["periodos_rango"] = periodos
    else:
        job_state["periodos_rango"] = None

    t = threading.Thread(target=run_automation, args=(usuario, password, periodo_principal, dl_path), daemon=True)
    t.start()
    return jsonify({"ok": True, "download_path": dl_path, "periodos_detectados": periodos})

@app.route("/api/stop", methods=["POST"])
def stop_job_route():
    with job_lock:
        if not job_state["running"]:
            return jsonify({"ok": False, "message": "No hay proceso en ejecución"}), 400
    stop_job()
    return jsonify({"ok": True, "message": "Deteniendo proceso..."})

@app.route("/api/reset", methods=["POST"])
def reset_job_route():
    data = request.json or {}
    periodo = data.get("periodo", "").strip()

    with job_lock:
        if job_state["running"]:
            stop_job()
            time.sleep(2)

    if periodo:
        periodo_dir = DOWNLOAD_DIR / periodo
        if periodo_dir.exists():
            for progreso_file in periodo_dir.glob("*/progreso.json"):
                try:
                    progreso_file.unlink()
                    log(f"🗑️ Progreso eliminado: {progreso_file}")
                except Exception as e:
                    log(f"⚠️ Error al borrar {progreso_file}: {e}", "warn")
        else:
            log(f"⚠️ No existe la carpeta del período '{periodo}'.", "warn")
    else:
        log("⚠️ No se especificó período, no se borró progreso.", "warn")

    reset_state()
    return jsonify({"ok": True, "message": "Estado reiniciado y progreso eliminado."})

@app.route("/api/status")
def get_status():
    with job_lock:
        return jsonify({
            "running": job_state["running"],
            "finished": job_state["finished"],
            "error": job_state["error"],
            "stats": job_state["stats"],
            "logs": job_state["logs"][-200:],
        })

@app.route("/api/logs")
def get_logs():
    since = int(request.args.get("since", 0))
    with job_lock:
        return jsonify({"logs": job_state["logs"][since:]})

@app.route("/api/logs", methods=["DELETE"])
def clear_logs():
    with job_lock:
        job_state["logs"] = []
    return jsonify({"ok": True})

@app.route("/api/files")
def list_files():
    periodo = request.args.get("periodo", "")
    folder = DOWNLOAD_DIR / periodo if periodo else DOWNLOAD_DIR
    files = []
    if folder.exists():
        for f in sorted(folder.iterdir()):
            if f.is_file():
                files.append({"name": f.name, "size": f.stat().st_size, "path": str(f), "periodo": periodo})
    return jsonify({"files": files})

@app.route("/api/files", methods=["DELETE"])
def delete_all_files():
    import shutil
    periodo = request.args.get("periodo", "")
    folder = DOWNLOAD_DIR / periodo if periodo else DOWNLOAD_DIR
    if not folder.exists():
        return jsonify({"ok": True, "message": "No hay archivos que eliminar"})
    try:
        eliminados = 0
        for item in list(folder.iterdir()):
            if item.is_file() and item.name != "progreso.json":
                item.unlink()
                eliminados += 1
            elif item.is_dir():
                # Dentro de subcarpetas: borrar archivos pero conservar progreso.json
                for sub in list(item.iterdir()):
                    if sub.is_file() and sub.name != "progreso.json":
                        sub.unlink()
                        eliminados += 1
                # Si la subcarpeta quedó vacía (o solo tiene progreso), dejarla
        log(f"🗑️ Soportes eliminados: {eliminados} archivo(s) en '{folder}' (progreso conservado)")
        return jsonify({"ok": True, "message": f"Se eliminaron {eliminados} soporte(s). El progreso se conservó.", "eliminados": eliminados})
    except Exception as e:
        log(f"⚠️ Error al eliminar soportes: {e}", "error")
        return jsonify({"ok": False, "error": str(e)}), 500

@app.route("/downloads/<path:filename>")
def download_file(filename):
    return send_from_directory(DOWNLOAD_DIR, filename, as_attachment=True)

@app.route("/api/periodos")
def get_periodos():
    periodos = []
    for d in DOWNLOAD_DIR.iterdir():
        if d.is_dir():
            count = len(list(d.glob("**/*.pdf")))
            periodos.append({"name": d.name, "count": count})
    return jsonify({"periodos": sorted(periodos, key=lambda x: x["name"], reverse=True)})

@app.route("/api/upload", methods=["POST"])
def upload_facturas():
    if 'file' not in request.files:
        return jsonify({"ok": False, "error": "No se envió ningún archivo"}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({"ok": False, "error": "Archivo vacío"}), 400

    try:
        filename = file.filename.lower()
        facturas = []
        if filename.endswith('.csv'):
            raw = file.read()
            # Detectar encoding: utf-8-sig cubre BOM, latin-1 cubre Windows
            for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
                try:
                    csv_text = raw.decode(enc)
                    break
                except (UnicodeDecodeError, LookupError):
                    continue
            else:
                csv_text = raw.decode('latin-1', errors='replace')
            reader = csv.DictReader(csv_text.splitlines())
            for row in reader:
                for col, val in row.items():
                    if 'factura' in col.lower():
                        facturas.append(val.strip())
                        break
        elif filename.endswith(('.xls', '.xlsx')):
            if not EXCEL_AVAILABLE:
                return jsonify({"ok": False, "error": "openpyxl no instalado"}), 500
            wb = openpyxl.load_workbook(BytesIO(file.read()), data_only=True)
            ws = wb.active
            col_idx = None
            for cell in ws[1]:
                if cell.value and 'factura' in str(cell.value).lower():
                    col_idx = cell.column
                    break
            if col_idx is None:
                return jsonify({"ok": False, "error": "No se encontró columna con 'factura'"}), 400
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[col_idx-1]
                if val:
                    facturas.append(str(val).strip())
        else:
            return jsonify({"ok": False, "error": "Formato no soportado. Use CSV o Excel"}), 400

        facturas_limpias = [re.sub(r'\D', '', f) for f in facturas if re.sub(r'\D', '', f)]
        if not facturas_limpias:
            return jsonify({"ok": False, "error": "No se encontraron números de factura válidos"}), 400

        with job_lock:
            job_state["facturas_permitidas"] = facturas_limpias
        log(f"📄 Se cargaron {len(facturas_limpias)} facturas desde el archivo.")
        return jsonify({"ok": True, "count": len(facturas_limpias), "facturas": facturas_limpias[:10]})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al procesar archivo: {str(e)}"}), 500

@app.route("/api/progreso")
def get_progreso():
    periodo = request.args.get("periodo", "")
    ips = request.args.get("ips", "")
    if not periodo:
        return jsonify({"ok": False, "error": "Se requiere el parámetro 'periodo'"}), 400

    periodo_dir = DOWNLOAD_DIR / periodo
    if not periodo_dir.exists():
        return jsonify({"ok": True, "completadas": [], "mensaje": "No hay datos para este período"})

    if ips:
        ips_dir = periodo_dir / ips
        if not ips_dir.exists():
            return jsonify({"ok": False, "error": f"No existe la IPS '{ips}'"}), 404
    else:
        posibles = list(periodo_dir.iterdir())
        if not posibles:
            return jsonify({"ok": True, "completadas": [], "mensaje": "No hay subcarpetas de IPS"})
        ips_dir = None
        for d in posibles:
            if d.is_dir() and (d / "progreso.json").exists():
                ips_dir = d
                break
        if not ips_dir:
            ips_dir = posibles[0] if posibles[0].is_dir() else None
        if not ips_dir:
            return jsonify({"ok": True, "completadas": [], "mensaje": "No se encontró carpeta de IPS"})
        ips = ips_dir.name

    progreso_path = ips_dir / "progreso.json"
    if not progreso_path.exists():
        return jsonify({"ok": True, "completadas": [], "ips": ips, "mensaje": "Aún no hay facturas completadas"})

    try:
        with open(progreso_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        completadas = data.get("completadas", [])
        return jsonify({"ok": True, "completadas": completadas, "cantidad": len(completadas), "ips": ips, "actualizado": data.get("actualizado", "")})
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al leer progreso: {str(e)}"}), 500

@app.route("/api/exportar_progreso")
def exportar_progreso_excel():
    periodo = request.args.get("periodo", "")
    if not periodo:
        return jsonify({"ok": False, "error": "Se requiere el parámetro 'periodo'"}), 400
    if not EXCEL_AVAILABLE:
        return jsonify({"ok": False, "error": "openpyxl no instalado"}), 500
    periodo_dir = DOWNLOAD_DIR / periodo
    if not periodo_dir.exists():
        return jsonify({"ok": False, "error": f"No existe la carpeta del período '{periodo}'"}), 404
    progreso_files = list(periodo_dir.glob("*/progreso.json"))
    if not progreso_files:
        return jsonify({"ok": False, "error": f"No se encontró progreso.json para el período '{periodo}'"}), 404
    progreso_path = progreso_files[0]
    ips_nombre = progreso_path.parent.name
    try:
        with open(progreso_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        completadas = data.get("completadas", [])
        actualizado = data.get("actualizado", "")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas completadas"
        ws.append(["N° Factura", "Fecha de completado"])
        for factura in completadas:
            ws.append([factura, actualizado])
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"progreso_facturas_{periodo}_{ips_nombre}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except Exception as e:
        return jsonify({"ok": False, "error": f"Error al generar Excel: {str(e)}"}), 500

if __name__ == "__main__":
    print("\n" + "=" * 55)
    print("  🏥 Activa IT — Descargador de Cartas Glosa")
    print("  🔷 Previsora SOAT (con nombres de IPS forzados desde el mapa)")
    print("=" * 55)
    print(f"  📂 Carpeta de descargas: {DOWNLOAD_DIR}")
    print(f"  🌐 Puerto: {port}")
    print("=" * 55 + "\n")
    app.run(host="0.0.0.0", port=port, debug=False)